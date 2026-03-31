from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone

def export_orders_to_excel(orders, user=None):
    """
    Export orders to Excel file with proper formatting
    
    Args:
        orders: QuerySet of Order objects
        user: Current user (for filename)
    
    Returns:
        HttpResponse with Excel file
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Define styles
    header_font = Font(name='Inter', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    cell_font = Font(name='Inter', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell_center = Alignment(horizontal='center', vertical='center')
    
    border_style = Side(border_style='thin', color='E5E7EB')
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # Headers
    headers = [
        'Order Number',
        'Customer Name',
        'Customer Phone',
        'Sales Person',
        'Items Count',
        'Total Quantity',
        'Order Date',
        'Notes'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    row_num = 2
    for order in orders:
        ws.cell(row=row_num, column=1, value=order.order_number)
        ws.cell(row=row_num, column=2, value=order.customer_name)
        ws.cell(row=row_num, column=3, value=order.customer_phone)
        ws.cell(row=row_num, column=4, value=order.sales_user.full_name if order.sales_user else 'N/A')
        ws.cell(row=row_num, column=5, value=order.total_unique_items)
        ws.cell(row=row_num, column=6, value=order.total_items)
        ws.cell(row=row_num, column=7, value=order.created_at.strftime('%d %b %Y, %H:%M'))
        ws.cell(row=row_num, column=8, value=order.notes or '-')
        
        # Apply styles to data rows
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.border = border
            
            # Center align for numeric columns
            if col_num in [5, 6]:
                cell.alignment = cell_center
            else:
                cell.alignment = cell_alignment
        
        row_num += 1
    
    # Adjust column widths
    column_widths = {
        'A': 20,  # Order Number
        'B': 25,  # Customer Name
        'C': 18,  # Phone
        'D': 20,  # Sales Person
        'E': 12,  # Items Count
        'F': 12,  # Total Qty
        'G': 20,  # Date
        'H': 30,  # Notes
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Generate filename
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    if user and user.is_admin:
        filename = f'All_Orders_{timestamp}.xlsx'
    else:
        filename = f'My_Orders_{timestamp}.xlsx'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


def export_order_details_to_excel(order):
    """
    Export single order with items to Excel
    
    Args:
        order: Order object
    
    Returns:
        HttpResponse with Excel file
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Details"
    
    # Styles
    header_font = Font(name='Inter', size=12, bold=True)
    label_font = Font(name='Inter', size=10, bold=True)
    value_font = Font(name='Inter', size=10)
    
    # Order Information Section
    ws.cell(row=1, column=1, value='ORDER INFORMATION').font = Font(name='Inter', size=14, bold=True)
    ws.merge_cells('A1:B1')
    
    row = 3
    info_fields = [
        ('Order Number:', order.order_number),
        ('Customer Name:', order.customer_name),
        ('Customer Phone:', order.customer_phone),
        ('Sales Person:', order.sales_user.full_name),
        ('Order Date:', order.created_at.strftime('%d %B %Y, %H:%M')),
        ('Notes:', order.notes or '-'),
    ]
    
    for label, value in info_fields:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value).font = value_font
        row += 1
    
    # Order Items Section
    row += 2
    ws.cell(row=row, column=1, value='ORDER ITEMS').font = Font(name='Inter', size=14, bold=True)
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 2
    item_headers = ['No', 'Merchandise', 'Category', 'Quantity']
    for col_num, header in enumerate(item_headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        cell.font = Font(name='Inter', size=11, bold=True, color='FFFFFF')
    
    row += 1
    for idx, item in enumerate(order.items.all(), 1):
        ws.cell(row=row, column=1, value=idx).font = value_font
        ws.cell(row=row, column=2, value=item.merchandise_name).font = value_font
        ws.cell(row=row, column=3, value=item.merchandise.category.name).font = value_font
        ws.cell(row=row, column=4, value=item.quantity).font = value_font
        row += 1
    
    # Total
    ws.cell(row=row, column=3, value='TOTAL:').font = label_font
    ws.cell(row=row, column=4, value=order.total_items).font = Font(name='Inter', size=10, bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    
    # Generate response
    filename = f'Order_{order.order_number}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
