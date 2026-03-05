from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from accounts.decorators import admin_required, admin_or_sales_required
from .models import Category, Merchandise, StockHistory
from .forms import CategoryForm, MerchandiseForm, StockAdjustmentForm, StockOpnameExportForm
import calendar
from datetime import datetime
from dateutil.relativedelta import relativedelta
from orders.models import OrderItem

# ============================================
# CATEGORY VIEWS
# ============================================

@admin_required
def category_list(request):
    """List all categories (Admin only)"""
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    categories = Category.objects.all()
    
    # Apply search filter
    if search_query:
        categories = categories.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Apply status filter
    if status_filter == 'active':
        categories = categories.filter(is_active=True)
    elif status_filter == 'inactive':
        categories = categories.filter(is_active=False)
    
    context = {
        'categories': categories,
        'search_query': search_query,
        'status_filter': status_filter
    }
    
    return render(request, 'merchandise/category_list.html', context)

@admin_required
def category_create(request):
    """Create new category (Admin only)"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            try:
                category = form.save()
                messages.success(request, f'Category "{category.name}" created successfully.')
                return redirect('merchandise:category_list')
            except Exception as e:
                messages.error(request, f'Error creating category: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CategoryForm()
    
    return render(request, 'merchandise/category_form.html', {
        'form': form,
        'title': 'Tambah Kategori',
        'button_text': 'Simpan Kategori'
    })

@admin_required
def category_edit(request, pk):
    """Edit existing category (Admin only)"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            try:
                category = form.save()
                messages.success(request, f'Category "{category.name}" updated successfully.')
                return redirect('merchandise:category_list')
            except Exception as e:
                messages.error(request, f'Error updating category: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CategoryForm(instance=category)
    
    return render(request, 'merchandise/category_form.html', {
        'form': form,
        'category': category,
        'title': f'Edit Kategori: {category.name}',
        'button_text': 'Simpan Perubahan'
    })

@admin_required
def category_delete(request, pk):
    """Delete category (Admin only) - Soft delete"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        category_name = category.name
        try:
            # Check if has merchandise
            if category.merchandise_set.exists():
                messages.error(
                    request,
                    f'Cannot delete category "{category_name}" because it has merchandise. '
                    'Please move or delete merchandise first.'
                )
            else:
                category.delete()  # Soft delete
                messages.success(request, f'Category "{category_name}" deleted successfully.')
            return redirect('merchandise:category_list')
        except Exception as e:
            messages.error(request, f'Error deleting category: {str(e)}')
            return redirect('merchandise:category_list')
    
    return render(request, 'merchandise/category_confirm_delete.html', {
        'category': category,
        'merchandise_count': category.merchandise_count
    })

@admin_required
def category_toggle_active(request, pk):
    """Toggle category active status (Admin only)"""
    category = get_object_or_404(Category, pk=pk)
    
    category.is_active = not category.is_active
    category.save()
    
    status = 'activated' if category.is_active else 'deactivated'
    messages.success(request, f'Category "{category.name}" has been {status}.')
    
    return redirect('merchandise:category_list')

# ============================================
# MERCHANDISE VIEWS
# ============================================

@admin_or_sales_required
def merchandise_list(request):
    """List all merchandise (Admin: full access, Sales: view only active items)"""
    search_query = request.GET.get('search', '')
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    stock_filter = request.GET.get('stock', '')
    
    # Filter based on role
    if request.user.is_admin:
        merchandise = Merchandise.objects.select_related('category', 'created_by').all()
    else:
        # Sales can only see active merchandise with stock
        merchandise = Merchandise.objects.select_related('category', 'created_by').filter(
            is_active=True
        )
    
    # Apply search filter
    if search_query:
        merchandise = merchandise.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Apply category filter
    if category_filter:
        merchandise = merchandise.filter(category_id=category_filter)
    
    # Apply status filter
    if status_filter == 'active':
        merchandise = merchandise.filter(is_active=True)
    elif status_filter == 'inactive':
        merchandise = merchandise.filter(is_active=False)
    
    # Apply stock filter
    if stock_filter == 'low':
        merchandise = merchandise.filter(stock__lt=10, stock__gt=0)
    elif stock_filter == 'out':
        merchandise = merchandise.filter(stock=0)
    
    categories = Category.objects.filter(is_active=True)
    
    context = {
        'merchandise': merchandise,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'stock_filter': stock_filter
    }
    
    return render(request, 'merchandise/merchandise_list.html', context)

@admin_required
def merchandise_create(request):
    """Create new merchandise (Admin only)"""
    if request.method == 'POST':
        form = MerchandiseForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                merchandise = form.save(commit=False)
                merchandise.created_by = request.user
                merchandise.save()
                messages.success(request, f'Merchandise "{merchandise.name}" berhasil ditambahkan.')
                return redirect('merchandise:merchandise_list')
            except Exception as e:
                messages.error(request, f'Gagal menambahkan merchandise: {str(e)}')
        else:
            messages.error(request, 'Harap perbaiki kesalahan di bawah ini.')
    else:
        form = MerchandiseForm()
    
    return render(request, 'merchandise/merchandise_form.html', {
        'form': form,
        'title': 'Tambah Merchandise Baru',
        'button_text': 'Simpan Merchandise'
    })

@admin_required
def merchandise_edit(request, pk):
    """Edit existing merchandise (Admin only)"""
    merchandise = get_object_or_404(Merchandise, pk=pk)
    
    if request.method == 'POST':
        form = MerchandiseForm(request.POST, request.FILES, instance=merchandise)
        if form.is_valid():
            try:
                merchandise = form.save()
                messages.success(request, f'Merchandise "{merchandise.name}" berhasil diperbarui.')
                return redirect('merchandise:merchandise_list')
            except Exception as e:
                messages.error(request, f'Gagal memperbarui merchandise: {str(e)}')
        else:
            messages.error(request, 'Harap perbaiki kesalahan di bawah ini.')
    else:
        form = MerchandiseForm(instance=merchandise)
    
    return render(request, 'merchandise/merchandise_form.html', {
        'form': form,
        'merchandise': merchandise,
        'title': f'Ubah Merchandise: {merchandise.name}',
        'button_text': 'Perbarui Merchandise'
    })

@admin_or_sales_required
def merchandise_detail(request, pk):
    """View merchandise details (Admin: full, Sales: view only)"""
    merchandise = get_object_or_404(
        Merchandise.objects.select_related('category', 'created_by'),
        pk=pk
    )
    stock_history = merchandise.stock_history.select_related('adjusted_by').all()[:10]
    
    context = {
        'merchandise': merchandise,
        'stock_history': stock_history
    }
    
    return render(request, 'merchandise/merchandise_detail.html', context)

@admin_required
def merchandise_delete(request, pk):
    """Delete merchandise (Admin only) - Soft delete"""
    merchandise = get_object_or_404(Merchandise, pk=pk)
    
    if request.method == 'POST':
        merchandise_name = merchandise.name
        try:
            merchandise.delete()  # Soft delete
            messages.success(request, f'Merchandise "{merchandise_name}" deleted successfully.')
            return redirect('merchandise:merchandise_list')
        except Exception as e:
            messages.error(request, f'Error deleting merchandise: {str(e)}')
            return redirect('merchandise:merchandise_list')
    
    return render(request, 'merchandise/merchandise_confirm_delete.html', {
        'merchandise': merchandise
    })

@admin_required
def merchandise_toggle_active(request, pk):
    """Toggle merchandise active status (Admin only)"""
    merchandise = get_object_or_404(Merchandise, pk=pk)
    
    merchandise.is_active = not merchandise.is_active
    merchandise.save()
    
    status = 'activated' if merchandise.is_active else 'deactivated'
    messages.success(request, f'Merchandise "{merchandise.name}" has been {status}.')
    
    return redirect('merchandise:merchandise_list')

@admin_required
def merchandise_adjust_stock(request, pk):
    """Adjust merchandise stock (Admin only)"""
    merchandise = get_object_or_404(Merchandise, pk=pk)

    # Block adjust stock untuk produk unlimited
    if merchandise.is_unlimited:
        messages.warning(
            request,
            f'"{merchandise.name}" memiliki stok unlimited — penyesuaian stok tidak diperlukan.'
        )
        return redirect('merchandise:merchandise_detail', pk=pk)

    if request.method == 'POST':
        form = StockAdjustmentForm(request.POST)
        if form.is_valid():
            try:
                adjustment = form.cleaned_data['adjustment']
                reason = form.cleaned_data['reason']

                # Create stock adjustment dalam atomic block untuk konsistensi
                with transaction.atomic():
                    StockHistory.create_adjustment(
                        merchandise=merchandise,
                        adjustment=adjustment,
                        reason=reason,
                        adjusted_by=request.user
                    )

                sign = '+' if adjustment >= 0 else ''
                messages.success(
                    request,
                    f'Stock adjusted: {sign}{adjustment}. New stock: {merchandise.stock}'
                )
                return redirect('merchandise:merchandise_detail', pk=pk)
            except Exception as e:
                messages.error(request, f'Error adjusting stock: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StockAdjustmentForm()
    
    return render(request, 'merchandise/stock_adjust.html', {
        'form': form,
        'merchandise': merchandise
    })

# ============================================
# STOCK OPNAME EXPORT
# ============================================

@admin_required
def stock_opname_export(request):
    """Export stock opname report to Excel (Admin only)"""
    if request.method == 'POST':
        form = StockOpnameExportForm(request.POST)
        if form.is_valid():
            try:
                month = int(form.cleaned_data['month'])
                year = int(form.cleaned_data['year'])
                branch_name = form.cleaned_data['branch_name']
                include_sales_tools = form.cleaned_data.get('include_sales_tools', True)
                
                # Generate Excel file
                response = generate_stock_opname_excel(month, year, branch_name, include_sales_tools)
                return response
            except Exception as e:
                messages.error(request, f'Error generating export: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StockOpnameExportForm()
    
    return render(request, 'merchandise/stock_opname_export.html', {
        'form': form,
        'title': 'Export Stock Opname'
    })


def generate_stock_opname_excel(month, year, branch_name, include_sales_tools=True):
    """Generate stock opname Excel file including merchandise and optionally sales tools"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum
    from django.utils import timezone
    from inventory.sales_tool_models import SalesTool, ToolCheckout
    
    # Calculate date range for the period (timezone-aware)
    start_date = timezone.make_aware(datetime(year, month, 1))
    end_date = start_date + relativedelta(months=1)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    month_name = calendar.month_name[month].upper()
    ws.title = f"STOCK OPNAME {month_name}"
    
    # Define styles
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    category_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=11)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title (Row 1-2 merged)
    ws.merge_cells('A1:D2')
    title_cell = ws['A1']
    title_cell.value = f'LAPORAN STOCK OPNAME - {month_name} {year}\n{branch_name}'
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[4].height = 25
    
    # Headers (Row 4)
    headers = ['JENIS MATERIAL/PRODUK', 'STOCK AWAL', 'PENGGUNAAN', 'SISA STOCK']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Get data - group by category
    current_row = 5
    categories = Category.objects.filter(is_active=True).prefetch_related('merchandise_set')
    
    for category in categories:
        merchandises = category.merchandise_set.filter(is_active=True)
        
        if not merchandises.exists():
            continue
        
        # Category row
        category_cell = ws.cell(row=current_row, column=1)
        category_cell.value = category.name.upper()
        category_cell.font = category_font
        category_cell.fill = category_fill
        category_cell.alignment = left_alignment
        category_cell.border = thin_border
        
        # Merge category row across columns
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        # Merchandise rows
        for merch in merchandises:
            # Calculate usage in period
            usage = OrderItem.objects.filter(
                merchandise=merch,
                order__created_at__gte=start_date,
                order__created_at__lte=end_date
            ).aggregate(total=Sum('quantity'))['total'] or 0

            # Product name
            name_cell = ws.cell(row=current_row, column=1)
            name_cell.font = normal_font
            name_cell.alignment = left_alignment
            name_cell.border = thin_border

            if merch.is_unlimited:
                # Unlimited: stok tidak pernah berkurang, tampilkan ∞
                unlimited_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                unlimited_font = Font(name='Arial', size=11, bold=True, color='7F6000')

                name_cell.value = f"  {merch.name}  [∞ UNLIMITED]"
                name_cell.fill = unlimited_fill

                stock_cell = ws.cell(row=current_row, column=2)
                stock_cell.value = "∞"
                stock_cell.font = unlimited_font
                stock_cell.alignment = center_alignment
                stock_cell.border = thin_border
                stock_cell.fill = unlimited_fill

                usage_cell = ws.cell(row=current_row, column=3)
                usage_cell.value = usage  # berapa kali dipakai
                usage_cell.font = normal_font
                usage_cell.alignment = center_alignment
                usage_cell.border = thin_border
                usage_cell.fill = unlimited_fill

                sisa_cell = ws.cell(row=current_row, column=4)
                sisa_cell.value = "∞"
                sisa_cell.font = unlimited_font
                sisa_cell.alignment = center_alignment
                sisa_cell.border = thin_border
                sisa_cell.fill = unlimited_fill
            else:
                # Stock awal = current stock + usage
                stock_awal = merch.stock + usage

                name_cell.value = f"  {merch.name}"

                stock_cell = ws.cell(row=current_row, column=2)
                stock_cell.value = stock_awal
                stock_cell.font = normal_font
                stock_cell.alignment = center_alignment
                stock_cell.border = thin_border

                usage_cell = ws.cell(row=current_row, column=3)
                usage_cell.value = usage
                usage_cell.font = normal_font
                usage_cell.alignment = center_alignment
                usage_cell.border = thin_border

                sisa_cell = ws.cell(row=current_row, column=4)
                sisa_cell.value = f'=B{current_row}-C{current_row}'
                sisa_cell.font = Font(name='Arial', size=11, bold=True)
                sisa_cell.alignment = center_alignment
                sisa_cell.border = thin_border

            current_row += 1
    
    # ============================================
    # SALES TOOLS SECTION
    # ============================================
    if include_sales_tools:
        # Add spacing row
        current_row += 1
        
        # Sales Tools Category Header
        tools_category_cell = ws.cell(row=current_row, column=1)
        tools_category_cell.value = "SALES TOOLS"
        tools_category_cell.font = category_font
        tools_category_cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        tools_category_cell.alignment = left_alignment
        tools_category_cell.border = thin_border
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        # Get active sales tools
        sales_tools = SalesTool.objects.filter(is_active=True).order_by('name')
        
        for tool in sales_tools:
            # Calculate usage in period (approved checkouts)
            usage = ToolCheckout.objects.filter(
                tool=tool,
                status='APPROVED',
                reviewed_at__gte=start_date,
                reviewed_at__lt=end_date
            ).aggregate(total=Sum('quantity'))['total'] or 0

            # Tool name
            name_cell = ws.cell(row=current_row, column=1)
            name_cell.font = normal_font
            name_cell.alignment = left_alignment
            name_cell.border = thin_border

            if tool.is_unlimited:
                unlimited_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                unlimited_font = Font(name='Arial', size=11, bold=True, color='7F6000')

                name_cell.value = f"  {tool.name}  [∞ UNLIMITED]"
                name_cell.fill = unlimited_fill

                stock_cell = ws.cell(row=current_row, column=2)
                stock_cell.value = "∞"
                stock_cell.font = unlimited_font
                stock_cell.alignment = center_alignment
                stock_cell.border = thin_border
                stock_cell.fill = unlimited_fill

                usage_cell = ws.cell(row=current_row, column=3)
                usage_cell.value = usage  # berapa kali checkout diapprove
                usage_cell.font = normal_font
                usage_cell.alignment = center_alignment
                usage_cell.border = thin_border
                usage_cell.fill = unlimited_fill

                sisa_cell = ws.cell(row=current_row, column=4)
                sisa_cell.value = "∞"
                sisa_cell.font = unlimited_font
                sisa_cell.alignment = center_alignment
                sisa_cell.border = thin_border
                sisa_cell.fill = unlimited_fill
            else:
                # Stock awal = current stock + usage
                stock_awal = tool.stock + usage

                name_cell.value = f"  {tool.name}"

                stock_cell = ws.cell(row=current_row, column=2)
                stock_cell.value = stock_awal
                stock_cell.font = normal_font
                stock_cell.alignment = center_alignment
                stock_cell.border = thin_border

                usage_cell = ws.cell(row=current_row, column=3)
                usage_cell.value = usage
                usage_cell.font = normal_font
                usage_cell.alignment = center_alignment
                usage_cell.border = thin_border

                sisa_cell = ws.cell(row=current_row, column=4)
                sisa_cell.value = f'=B{current_row}-C{current_row}'
                sisa_cell.font = Font(name='Arial', size=11, bold=True)
                sisa_cell.alignment = center_alignment
                sisa_cell.border = thin_border

            current_row += 1
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Stock_Opname_{month_name}_{year}_{branch_name}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
